#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Robô DataJud / CNJ — monitoramento nacional de distribuições
Classes: 129 (Recuperação Judicial), 128 (Recuperação Extrajudicial),
         108 (Falência de Empresários, Sociedades Empresariais, ME e EPP)

Roda em GitHub Actions (Ubuntu) e grava os resultados no próprio repositório.
"""

import os
import sys
import json
import time
import datetime
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------
# CONFIGURAÇÃO
# ----------------------------------------------------------------------

# --- Google Drive (opcional; ativo quando as duas variáveis existem) ---
# GDRIVE_CREDENTIALS_JSON : conteúdo do JSON da conta de serviço (Secret)
# GDRIVE_FOLDER_ID        : ID da pasta de destino no Drive (Secret)
GDRIVE_CREDENTIALS_JSON = os.environ.get("GDRIVE_CREDENTIALS_JSON", "")
GDRIVE_FOLDER_ID = os.environ.get("GDRIVE_FOLDER_ID", "")

# Chave pública oficial do CNJ (divulgada em datajud-wiki.cnj.jus.br/api-publica/acesso).
# Pode ser sobrescrita por Secret do repositório caso o CNJ a altere.
API_KEY = os.environ.get(
    "DATAJUD_API_KEY",
    "cDZHYzlZa0JadVREZDJCendQbXY6SkJlTzNjLV9TRENyQk1RdnFKZGRQdw==",
)

# Janela de ajuizamento considerada, em dias (a base tem 5-7 semanas de atraso).
DIAS_JANELA = int(os.environ.get("DIAS_JANELA", "365"))

CLASSES_CNJ = {
    129: "Recuperação Judicial",
    128: "Recuperação Extrajudicial",
    108: "Falência de Empresários/Sociedades/ME/EPP",
}

# Tribunais monitorados. Pode ser sobrescrito por variável de ambiente,
# ex.: TRIBUNAIS="tjsp,tjrj"  (siglas separadas por vírgula)
TRIBUNAIS = [
    s.strip().lower()
    for s in os.environ.get("TRIBUNAIS", "tjsp,tjrj,tjpr,tjmg,tjrs").split(",")
    if s.strip()
]

BASE_URL = "https://api-publica.datajud.cnj.jus.br/api_publica_{}/_search"
PASTA_SAIDA = os.environ.get("PASTA_SAIDA", "saida")
PASTA_ESTADO = os.environ.get("PASTA_ESTADO", "estado")
ARQ_ESTADO = os.path.join(PASTA_ESTADO, "processos_vistos.json")
PAUSA = 1.0          # segundos entre requisições (cortesia com a API pública)
PAGINA = 500         # registros por página
TIMEOUT = 45


def log(msg):
    print(msg, flush=True)


def formatar_processo(num):
    """20 dígitos -> NNNNNNN-DD.AAAA.J.TR.OOOO"""
    n = "".join(c for c in str(num) if c.isdigit())
    if len(n) != 20:
        return str(num)
    return f"{n[0:7]}-{n[7:9]}.{n[9:13]}.{n[13]}.{n[14:16]}.{n[16:20]}"


def formatar_data(bruto):
    """DataJud devolve 'AAAAMMDDHHMMSS' (string), não ISO-8601."""
    s = str(bruto or "").strip()
    for fmt in ("%Y%m%d%H%M%S", "%Y-%m-%dT%H:%M:%S.%fZ", "%Y-%m-%dT%H:%M:%SZ"):
        try:
            return datetime.datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def consultar_tribunal(sigla, data_corte):
    """Busca todas as distribuições do tribunal na janela, com paginação."""
    url = BASE_URL.format(sigla)
    headers = {"Authorization": f"APIKey {API_KEY}", "Content-Type": "application/json"}
    coletados, search_after = [], None

    while True:
        payload = {
            "query": {"bool": {"filter": [
                {"terms": {"classe.codigo": list(CLASSES_CNJ.keys())}},
                {"range": {"dataAjuizamento": {"gte": data_corte}}},
            ]}},
            "sort": [{"dataAjuizamento": {"order": "desc"}},
                     {"id.keyword": {"order": "asc"}}],
            "size": PAGINA,
        }
        if search_after:
            payload["search_after"] = search_after

        try:
            r = requests.post(url, json=payload, headers=headers, timeout=TIMEOUT)
        except Exception as e:
            log(f"    !! {sigla.upper()}: falha de rede ({type(e).__name__})")
            return coletados, False

        # ---- tratamento explícito de erro: 401/403/429 NÃO podem passar em silêncio
        if r.status_code != 200:
            log(f"    !! {sigla.upper()}: HTTP {r.status_code} — {r.text[:180]}")
            return coletados, False

        hits = r.json().get("hits", {}).get("hits", [])
        if not hits:
            break

        for h in hits:
            s = h.get("_source", {})
            dt_aj = formatar_data(s.get("dataAjuizamento"))
            dt_up = formatar_data(s.get("dataHoraUltimaAtualizacao"))
            classe = s.get("classe", {}) or {}
            assuntos = s.get("assuntos", []) or []
            coletados.append({
                "id": s.get("id") or h.get("_id"),
                "tribunal": s.get("tribunal", sigla.upper()),
                "grau": s.get("grau", ""),
                "numero": formatar_processo(s.get("numeroProcesso", "")),
                "classe": CLASSES_CNJ.get(classe.get("codigo"), classe.get("nome", "")),
                "data_aj": dt_aj,
                "orgao": (s.get("orgaoJulgador", {}) or {}).get("nome", ""),
                "municipio": (s.get("orgaoJulgador", {}) or {}).get("codigoMunicipioIBGE", ""),
                "assunto": assuntos[0].get("nome", "") if assuntos else "",
                "atualizado": dt_up,
            })

        search_after = hits[-1].get("sort")
        if len(hits) < PAGINA or not search_after:
            break
        time.sleep(PAUSA)

    return coletados, True


def gerar_planilha(registros, caminho, data_exec):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Distribuições"

    cabecalho = [
        "Status", "Tribunal", "Grau", "Número do Processo", "Classe Processual",
        "Data de Ajuizamento", "Assunto Principal", "Órgão Julgador",
        "Cód. Município IBGE", "Última Atualização (base CNJ)",
    ]
    ws.append(cabecalho)

    azul = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    verde = PatternFill("solid", start_color="E8F5E9", end_color="E8F5E9")
    zebra = PatternFill("solid", start_color="F2F4F8", end_color="F2F4F8")
    branco = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
    borda = Border(*[Side(style="thin", color="D9D9D9")] * 4)

    for c in range(1, len(cabecalho) + 1):
        cel = ws.cell(row=1, column=c)
        cel.fill = azul
        cel.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    for i, p in enumerate(registros, start=2):
        novo = p["status"] == "NOVO"
        valores = [
            p["status"], p["tribunal"], p["grau"], p["numero"], p["classe"],
            p["data_aj"], p["assunto"], p["orgao"], p["municipio"], p["atualizado"],
        ]
        for c, v in enumerate(valores, start=1):
            cel = ws.cell(row=i, column=c, value=v)
            cel.font = Font(name="Arial", size=9, bold=(novo and c == 1))
            cel.border = borda
            cel.fill = verde if novo else (zebra if i % 2 == 0 else branco)
            if c in (1, 2, 3, 6, 9, 10):
                cel.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cel.alignment = Alignment(horizontal="left", vertical="center")
            if c in (6, 10) and isinstance(v, datetime.datetime):
                cel.number_format = "DD/MM/YYYY HH:MM"

    ws.auto_filter.ref = f"A1:{get_column_letter(len(cabecalho))}{max(ws.max_row, 1)}"
    larguras = [10, 10, 7, 26, 34, 20, 34, 42, 16, 22]
    for c, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # Aba de resumo (contagens por tribunal, com fórmulas vivas)
    ws2 = wb.create_sheet("Resumo")
    ws2.append(["Execução", data_exec.strftime("%d/%m/%Y %H:%M")])
    ws2.append(["Janela (dias)", DIAS_JANELA])
    ws2.append([])
    ws2.append(["Tribunal", "Total", "Novos nesta execução"])
    ult = max(ws.max_row, 2)
    for j, t in enumerate(sorted({p["tribunal"] for p in registros}), start=5):
        ws2.cell(row=j, column=1, value=t)
        ws2.cell(row=j, column=2, value=f'=COUNTIF(Distribuições!$B$2:$B${ult},$A{j})')
        ws2.cell(row=j, column=3,
                 value=f'=COUNTIFS(Distribuições!$B$2:$B${ult},$A{j},Distribuições!$A$2:$A${ult},"NOVO")')
    for c in range(1, 4):
        ws2.cell(row=4, column=c).font = Font(name="Arial", size=10, bold=True)
        ws2.column_dimensions[get_column_letter(c)].width = 24

    os.makedirs(os.path.dirname(caminho) or ".", exist_ok=True)
    wb.save(caminho)


def enviar_para_drive(caminho_local, nome_arquivo):
    """
    Envia/atualiza a planilha numa pasta do Google Drive via conta de serviço.

    Importante: contas de serviço NÃO têm cota de armazenamento no "Meu Drive".
    Por isso a estratégia é ATUALIZAR um arquivo já existente na pasta
    (criado uma única vez pelo dono, que detém a cota). Só tenta criar um
    arquivo novo se nenhum for encontrado — o que funciona apenas em
    Drives Compartilhados.
    """
    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaFileUpload
    except ImportError:
        log("    !! Drive: bibliotecas google-api-python-client/google-auth ausentes.")
        return False

    try:
        info = json.loads(GDRIVE_CREDENTIALS_JSON)
        creds = service_account.Credentials.from_service_account_info(
            info, scopes=["https://www.googleapis.com/auth/drive"])
        svc = build("drive", "v3", credentials=creds, cache_discovery=False)

        # Procura o arquivo pelo nome dentro da pasta de destino
        q = (f"name = '{nome_arquivo}' and '{GDRIVE_FOLDER_ID}' in parents "
             f"and trashed = false")
        res = svc.files().list(q=q, fields="files(id, name)",
                               supportsAllDrives=True,
                               includeItemsFromAllDrives=True).execute()
        arquivos = res.get("files", [])

        media = MediaFileUpload(
            caminho_local,
            mimetype="application/vnd.openxmlformats-officedocument."
                     "spreadsheetml.sheet",
            resumable=True)

        if arquivos:
            fid = arquivos[0]["id"]
            svc.files().update(fileId=fid, media_body=media,
                               supportsAllDrives=True).execute()
            log(f"  Drive: planilha ATUALIZADA (id {fid[:12]}…)")
        else:
            meta = {"name": nome_arquivo, "parents": [GDRIVE_FOLDER_ID]}
            novo = svc.files().create(body=meta, media_body=media,
                                      fields="id",
                                      supportsAllDrives=True).execute()
            log(f"  Drive: planilha CRIADA (id {novo['id'][:12]}…)")
        return True

    except Exception as e:
        log(f"    !! Drive: falha no envio — {type(e).__name__}: {e}")
        log("       Se o erro citar 'storageQuotaExceeded', crie manualmente um")
        log("       arquivo com esse exato nome na pasta e rode de novo (o robô")
        log("       passará a atualizá-lo em vez de criar).")
        return False


def main():
    inicio = datetime.datetime.now()
    corte = (inicio - datetime.timedelta(days=DIAS_JANELA)).strftime("%Y%m%d%H%M%S")
    log(f"Robô DataJud/CNJ — início {inicio:%d/%m/%Y %H:%M}")
    log(f"Classes: {', '.join(str(c) for c in CLASSES_CNJ)} | janela: {DIAS_JANELA} dias\n")

    # estado anterior (ids já vistos em execuções passadas)
    vistos = set()
    if os.path.exists(ARQ_ESTADO):
        with open(ARQ_ESTADO, encoding="utf-8") as f:
            vistos = set(json.load(f).get("ids", []))
    log(f"Processos já conhecidos: {len(vistos)}\n")

    todos, falhas = [], []
    for sigla in TRIBUNAIS:
        regs, ok = consultar_tribunal(sigla, corte)
        if not ok:
            falhas.append(sigla.upper())
        novos = sum(1 for r in regs if r["id"] not in vistos)
        log(f"  {sigla.upper():<6} {len(regs):>4} processos ({novos} novos)")
        todos.extend(regs)
        time.sleep(PAUSA)

    # ---- se TODOS os tribunais falharam, é erro de configuração: aborta com exit 1
    if falhas and len(falhas) == len(TRIBUNAIS):
        log(f"\nERRO FATAL: todos os {len(falhas)} tribunais falharam.")
        log("Se o log acima mostra HTTP 401/403, a API Key está inválida ou foi")
        log("alterada pelo CNJ — confira a chave vigente em:")
        log("  https://datajud-wiki.cnj.jus.br/api-publica/acesso")
        log("Se mostra HTTP 400, a query está malformada. Se 429, reduza a cadência.")
        sys.exit(1)

    for r in todos:
        r["status"] = "NOVO" if r["id"] not in vistos else "Monitorado"
    todos.sort(key=lambda r: (r["status"] != "NOVO", r["data_aj"] or datetime.datetime.min),
               reverse=False)
    todos.sort(key=lambda r: (r["status"] != "NOVO",
                              -(r["data_aj"] or datetime.datetime.min).timestamp()))

    qtd_novos = sum(1 for r in todos if r["status"] == "NOVO")

    saida = os.path.join(PASTA_SAIDA, "Distribuicao_RJ_RE_Falencia.xlsx")
    gerar_planilha(todos, saida, inicio)

    os.makedirs(PASTA_ESTADO, exist_ok=True)
    with open(ARQ_ESTADO, "w", encoding="utf-8") as f:
        json.dump({"atualizado_em": inicio.isoformat(),
                   "ids": sorted({r["id"] for r in todos} | vistos)}, f,
                  ensure_ascii=False, indent=1)

    log(f"\n{'='*58}")
    log(f"Total coletado : {len(todos)}")
    log(f"Novos          : {qtd_novos}")
    if falhas:
        log(f"Tribunais c/ falha: {', '.join(falhas)}")
    log(f"Planilha       : {saida}")

    # --- envio ao Google Drive (se configurado) ---
    drive_ok = None
    if GDRIVE_CREDENTIALS_JSON and GDRIVE_FOLDER_ID:
        drive_ok = enviar_para_drive(saida, os.path.basename(saida))
    else:
        log("  Drive: não configurado (defina GDRIVE_CREDENTIALS_JSON e "
            "GDRIVE_FOLDER_ID)")

    log(f"Duração        : {(datetime.datetime.now()-inicio).seconds}s")

    # expõe o resultado para o resumo do GitHub Actions
    if os.environ.get("GITHUB_STEP_SUMMARY"):
        st_drive = ("enviado ✅" if drive_ok
                    else "FALHOU ❌" if drive_ok is False else "não configurado")
        with open(os.environ["GITHUB_STEP_SUMMARY"], "a", encoding="utf-8") as f:
            f.write(f"### Robô DataJud\n\n"
                    f"- Coletados: **{len(todos)}**\n- Novos: **{qtd_novos}**\n"
                    f"- Falhas: {', '.join(falhas) if falhas else 'nenhuma'}\n"
                    f"- Google Drive: {st_drive}\n")


if __name__ == "__main__":
    main()
